import os
import re
import json
import html
import time
import traceback
from datetime import datetime, timedelta
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


# ============================================================
# CONFIG
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(BASE_DIR, "tv360channels.xlsx")
XML_FILE = os.path.join(BASE_DIR, "tv360epg.xml")
LOG_FILE = os.path.join(BASE_DIR, "tv360log.txt")

CHANNEL_PAGE = "https://m.tv360.vn/tv/"
SCHEDULE_API = "https://m.tv360.vn/public/v1/live/get-live-schedule?id={}"

SOURCE_INFO_NAME = "Ngân Phúc"
SOURCE_INFO_URL = "https://epg.vercel.app/epg.xml"
GENERATOR_INFO_NAME = "EPG GitHub"

MAX_LOG_DAYS = 7

REQUEST_TIMEOUT = 30

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://m.tv360.vn/",
}


# ============================================================
# UTILS
# ============================================================

def log_print(message):
    print(message, flush=True)


def clean_text(value):
    if value is None:
        return ""

    value = str(value)

    # HTML entities
    value = html.unescape(value)

    # Không xuống dòng
    value = value.replace("\r", " ").replace("\n", " ").replace("\t", " ")

    # Xóa khoảng trắng trước , :
    value = re.sub(r"\s+([,:])", r"\1", value)

    # Thêm khoảng trắng sau , :
    value = re.sub(r"([,:])(?=\S)", r"\1 ", value)

    # Xóa dấu , hoặc : nếu ở cuối chuỗi
    value = re.sub(r"([,:])\s*$", "", value)

    # Dấu -
    value = re.sub(r"\s*-\s*", " - ", value)

    # Loại bỏ khoảng trắng kép
    value = re.sub(r"\s+", " ", value)

    return value.strip()


def xml_escape(value):
    """
    Escape XML theo đúng yêu cầu:
        &  -> &amp;
        <  -> &lt;
        >  -> &gt;
        "  -> &quot;
        '  -> &apos;
    """

    if value is None:
        return ""

    value = str(value)

    # Quan trọng: phải escape & trước
    value = value.replace("&", "&amp;")
    value = value.replace("<", "&lt;")
    value = value.replace(">", "&gt;")
    value = value.replace('"', "&quot;")
    value = value.replace("'", "&apos;")

    return value


def normalize_xml_text(value):
    return xml_escape(clean_text(value))


# ============================================================
# HTTP
# ============================================================

session = requests.Session()
session.headers.update(HEADERS)


def get_response(url):
    response = session.get(
        url,
        timeout=REQUEST_TIMEOUT
    )

    response.raise_for_status()

    return response


# ============================================================
# CHANNEL SCRAPING
# ============================================================

def recursive_find_channel_objects(obj, result=None):
    """
    Tìm các object có khả năng là channel trong JSON bất kể
    cấu trúc JSON của TV360 thay đổi nhẹ.

    Các field quan trọng:
        id
        name
        slug
        link
        coverImage
    """

    if result is None:
        result = []

    if isinstance(obj, dict):

        keys_lower = {
            str(k).lower(): k
            for k in obj.keys()
        }

        id_key = None
        name_key = None

        for candidate in ["id", "channel_id", "channelid", "liveid"]:
            if candidate in keys_lower:
                id_key = keys_lower[candidate]
                break

        for candidate in ["name", "displayname", "display-name", "title"]:
            if candidate in keys_lower:
                name_key = keys_lower[candidate]
                break

        if id_key is not None and name_key is not None:

            channel_id = obj.get(id_key)
            name = obj.get(name_key)

            if channel_id is not None and name:

                channel_id = str(channel_id).strip()
                name = str(name).strip()

                # Tránh nhận nhầm object không phải channel
                if channel_id and name:

                    result.append(obj)

        for value in obj.values():
            recursive_find_channel_objects(value, result)

    elif isinstance(obj, list):

        for item in obj:
            recursive_find_channel_objects(item, result)

    return result


def extract_channels_from_json_scripts(soup):
    """
    Tìm JSON nằm trong các <script> của trang.
    """

    channels = []

    for script in soup.find_all("script"):

        content = script.string or script.get_text()

        if not content:
            continue

        content = content.strip()

        # Chỉ thử JSON nguyên khối
        if content.startswith("{") or content.startswith("["):

            try:
                data = json.loads(content)

                found = recursive_find_channel_objects(data)

                channels.extend(found)

            except Exception:
                pass

    return channels


def extract_channels_from_html_links(soup):
    """
    Fallback:
    tìm các link dạng /tv/xxx?ch=194
    """

    channels = []

    for a in soup.find_all("a", href=True):

        href = a.get("href", "").strip()

        if not href:
            continue

        full_url = urljoin(CHANNEL_PAGE, href)

        match = re.search(
            r"(?:[?&]ch=)(\d+)",
            full_url,
            re.IGNORECASE
        )

        if not match:
            continue

        channel_id = match.group(1)

        text = a.get_text(" ", strip=True)

        if not text:
            continue

        channels.append({
            "id": channel_id,
            "name": text,
            "slug": "",
            "link": full_url,
            "coverImage": "",
        })

    return channels


def extract_channels_from_html_regex(html_text):
    """
    Fallback cuối cùng:
    tìm URL TV360 dạng:

        /tv/slug?ch=194
    """

    channels = []

    pattern = re.compile(
        r'["\']([^"\']*/tv/([^"\']+?)(?:\?|\&)ch=(\d+)[^"\']*)["\']',
        re.IGNORECASE
    )

    for match in pattern.finditer(html_text):

        link = match.group(1)
        slug = match.group(2)
        channel_id = match.group(3)

        link = urljoin(CHANNEL_PAGE, link)

        channels.append({
            "id": channel_id,
            "name": slug.replace("-", " "),
            "slug": slug,
            "link": link,
            "coverImage": "",
        })

    return channels


def normalize_channel_object(obj):
    """
    Chuẩn hóa object API thành:

        id
        name
        slug
        link
        coverImage
    """

    def get_value(keys):
        for key in keys:

            if key in obj:
                value = obj.get(key)

                if value is not None:
                    return value

        return ""

    channel_id = get_value([
        "id",
        "channel_id",
        "channelId",
        "liveId",
    ])

    name = get_value([
        "name",
        "displayName",
        "display-name",
        "title",
    ])

    slug = get_value([
        "slug",
        "channelSlug",
    ])

    link = get_value([
        "link",
        "url",
        "href",
    ])

    cover = get_value([
        "coverImage",
        "cover_image",
        "coverimage",
        "horizontalImage",
        "horizontal_image",
    ])

    if isinstance(channel_id, dict):
        channel_id = channel_id.get("id", "")

    channel_id = str(channel_id).strip()
    name = str(name).strip()
    slug = str(slug).strip()
    link = str(link).strip()
    cover = str(cover).strip()

    if link:
        link = urljoin(CHANNEL_PAGE, link)

    if cover:
        cover = urljoin(CHANNEL_PAGE, cover)

    if not link and slug and channel_id:
        link = f"https://tv360.vn/tv/{slug}?ch={channel_id}"

    return {
        "id": channel_id,
        "name": name,
        "slug": slug,
        "link": link,
        "coverImage": cover,
    }


def get_all_channels():
    """
    Lấy toàn bộ danh sách kênh TV360.

    Vì cấu trúc frontend TV360 có thể thay đổi,
    hàm này sử dụng nhiều lớp fallback.
    """

    log_print("Đang lấy danh sách kênh TV360...")

    response = get_response(CHANNEL_PAGE)

    html_text = response.text

    soup = BeautifulSoup(html_text, "html.parser")

    candidates = []

    # --------------------------------------------------------
    # 1. JSON trong script
    # --------------------------------------------------------

    candidates.extend(
        extract_channels_from_json_scripts(soup)
    )

    # --------------------------------------------------------
    # 2. Link HTML
    # --------------------------------------------------------

    candidates.extend(
        extract_channels_from_html_links(soup)
    )

    # --------------------------------------------------------
    # 3. Regex
    # --------------------------------------------------------

    candidates.extend(
        extract_channels_from_html_regex(html_text)
    )

    channels = {}

    for obj in candidates:

        channel = normalize_channel_object(obj)

        channel_id = channel["id"]

        if not channel_id:
            continue

        # Tên phải có
        if not channel["name"]:
            continue

        # Tránh trùng
        channels[channel_id] = channel

    result = list(channels.values())

    result.sort(
        key=lambda x: (
            int(x["id"]) if x["id"].isdigit() else 999999999,
            x["name"].lower()
        )
    )

    log_print(
        f"Tìm thấy {len(result)} kênh."
    )

    return result


# ============================================================
# EXCEL
# ============================================================

HEADERS_DATA = [
    "id",
    "name",
    "slug",
    "link",
    "coverImage",
    "channel",
    "display-name",
]


def ensure_workbook():
    """
    Nếu chưa có tv360channels.xlsx thì tạo mới.

    Nếu đã có:
        - giữ nguyên workbook
        - giữ nguyên Tham chiếu
        - giữ nguyên F:G
    """

    if os.path.exists(EXCEL_FILE):

        wb = load_workbook(
            EXCEL_FILE,
            data_only=False
        )

    else:

        wb = Workbook()

        ws = wb.active
        ws.title = "Data"

        ws_ref = wb.create_sheet("Tham chiếu")

        for col, header in enumerate(HEADERS_DATA, 1):

            ws.cell(
                row=1,
                column=col,
                value=header
            )

            ws_ref.cell(
                row=1,
                column=col,
                value=header
            )

            ws.cell(
                row=1,
                column=col
            ).font = Font(bold=True)

            ws_ref.cell(
                row=1,
                column=col
            ).font = Font(bold=True)

        wb.save(EXCEL_FILE)

    # --------------------------------------------------------
    # Đảm bảo 2 sheet tồn tại
    # --------------------------------------------------------

    if "Data" not in wb.sheetnames:
        ws = wb.create_sheet("Data")
    else:
        ws = wb["Data"]

    if "Tham chiếu" not in wb.sheetnames:
        ws_ref = wb.create_sheet("Tham chiếu")
    else:
        ws_ref = wb["Tham chiếu"]

    # --------------------------------------------------------
    # Đảm bảo header
    # --------------------------------------------------------

    for col, header in enumerate(HEADERS_DATA, 1):

        if ws.cell(1, col).value is None:
            ws.cell(1, col).value = header

        if ws_ref.cell(1, col).value is None:
            ws_ref.cell(1, col).value = header

    return wb, ws, ws_ref


def read_reference_mapping(ws_ref):
    """
    Đọc mapping thủ công từ:

        Tham chiếu!A = id
        Tham chiếu!F = channel
        Tham chiếu!G = display-name

    Đây mới là mapping được Python dùng để tạo XML.
    """

    mapping = {}

    for row in range(2, ws_ref.max_row + 1):

        channel_id = ws_ref.cell(row, 1).value

        if channel_id is None:
            continue

        channel_id = str(channel_id).strip()

        if not channel_id:
            continue

        channel = ws_ref.cell(row, 6).value
        display_name = ws_ref.cell(row, 7).value

        channel = (
            str(channel).strip()
            if channel is not None
            else ""
        )

        display_name = (
            str(display_name).strip()
            if display_name is not None
            else ""
        )

        mapping[channel_id] = {
            "channel": channel,
            "display-name": display_name,
        }

    return mapping


def write_channels_to_excel(wb, ws, channels):
    """
    Chỉ cập nhật A:E.

    Tuyệt đối không xóa F:G.

    Các dòng F:G cũ vẫn được giữ nguyên.
    """

    # --------------------------------------------------------
    # Xóa dữ liệu cũ A:E
    # --------------------------------------------------------

    if ws.max_row >= 2:

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=5
        ):
            for cell in row:
                cell.value = None

    # --------------------------------------------------------
    # Ghi dữ liệu mới A:E
    # --------------------------------------------------------

    for row_index, channel in enumerate(channels, 2):

        ws.cell(
            row=row_index,
            column=1,
            value=channel["id"]
        )

        ws.cell(
            row=row_index,
            column=2,
            value=channel["name"]
        )

        ws.cell(
            row=row_index,
            column=3,
            value=channel["slug"]
        )

        ws.cell(
            row=row_index,
            column=4,
            value=channel["link"]
        )

        ws.cell(
            row=row_index,
            column=5,
            value=channel["coverImage"]
        )

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    for col in range(1, 8):

        ws.cell(
            row=1,
            column=col
        ).font = Font(bold=True)

        ws.cell(
            row=1,
            column=col
        ).alignment = Alignment(
            horizontal="center"
        )

    # --------------------------------------------------------
    # Không tự động xóa F:G
    #
    # Chỉ đảm bảo formula nếu người dùng đã có formula.
    #
    # Nếu workbook mới hoàn toàn thì tạo VLOOKUP.
    # --------------------------------------------------------

    has_existing_formula = False

    for row in range(2, min(ws.max_row, 20) + 1):

        f = ws.cell(row, 6).value
        g = ws.cell(row, 7).value

        if (
            isinstance(f, str)
            and f.startswith("=")
        ) or (
            isinstance(g, str)
            and g.startswith("=")
        ):
            has_existing_formula = True
            break

    # Workbook mới:
    # tạo formula cho số dòng hiện tại + 20 dòng dự phòng.
    if not has_existing_formula:

        formula_end = max(
            len(channels) + 20,
            20
        )

        for row in range(2, formula_end + 1):

            ws.cell(
                row=row,
                column=6,
                value=(
                    f'=IFERROR(VLOOKUP(A{row},'
                    f"'Tham chiếu'!$A:$G,6,FALSE),\"\")"
                )
            )

            ws.cell(
                row=row,
                column=7,
                value=(
                    f'=IFERROR(VLOOKUP(A{row},'
                    f"'Tham chiếu'!$A:$G,7,FALSE),\"\")"
                )
            )

    # --------------------------------------------------------
    # Chiều rộng cột
    # --------------------------------------------------------

    widths = {
        "A": 14,
        "B": 35,
        "C": 35,
        "D": 60,
        "E": 70,
        "F": 20,
        "G": 35,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(EXCEL_FILE)


# ============================================================
# EPG
# ============================================================

def get_epg_for_channel(channel_id):
    """
    Lấy EPG cho một ID TV360.
    """

    url = SCHEDULE_API.format(channel_id)

    try:

        response = get_response(url)

        data = response.json()

        schedules = (
            data
            .get("data", {})
            .get("schedules", [])
        )

        if not isinstance(schedules, list):
            return []

        return schedules

    except Exception as exc:

        log_print(
            f"  Lỗi EPG ID {channel_id}: {exc}"
        )

        return []


def parse_datetime(date_str, time_str):
    """
    TV360:
        datetime = 2026-08-22
        startTime = 15:25

    Output:
        datetime object
    """

    try:

        return datetime.strptime(
            f"{date_str} {time_str}",
            "%Y-%m-%d %H:%M"
        )

    except Exception:

        return None


def format_xmltv_datetime(dt):
    """
    XMLTV:
        YYYYMMDDHHMMSS +0700
    """

    return dt.strftime(
        "%Y%m%d%H%M%S +0700"
    )


def duration_minutes(start_dt, end_dt):
    """
    Tính thời lượng phút.

    Làm tròn lên nếu có giây lẻ.
    """

    seconds = (
        end_dt - start_dt
    ).total_seconds()

    if seconds <= 0:
        return 0

    return int(
        (seconds + 59) // 60
    )


def build_programme(
    channel,
    display_name,
    schedule
):
    """
    Tạo một programme XML.
    """

    title = schedule.get("name", "")

    date_str = schedule.get(
        "datetime",
        ""
    )

    start_time = schedule.get(
        "startTime",
        ""
    )

    end_time = schedule.get(
        "endTime",
        ""
    )

    start_dt = parse_datetime(
        date_str,
        start_time
    )

    end_dt = parse_datetime(
        date_str,
        end_time
    )

    if not start_dt or not end_dt:
        return None

    # Trường hợp chương trình qua 00:00
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    minutes = duration_minutes(
        start_dt,
        end_dt
    )

    title = normalize_xml_text(title)

    return (
        f'  <programme '
        f'start="{format_xmltv_datetime(start_dt)}" '
        f'stop="{format_xmltv_datetime(end_dt)}" '
        f'channel="{xml_escape(channel)}">\n'
        f'    <title lang="vi">{title}</title>\n'
        f'    <length lang="vi">'
        f'Chương trình này có thời lượng {minutes} phút'
        f'</length>\n'
        f'  </programme>\n'
    )


def build_xml(
    channels,
    epg_data,
    mapping
):
    """
    Tạo toàn bộ tv360epg.xml.

    mapping:
        TV360 ID -> channel/display-name

    Chỉ những channel đã mapping mới được đưa vào XML.
    """

    lines = []

    lines.append(
        '<?xml version="1.0" encoding="UTF-8"?>'
    )

    lines.append(
        f'<tv source-info-name="{xml_escape(SOURCE_INFO_NAME)}" '
        f'source-info-url="{xml_escape(SOURCE_INFO_URL)}" '
        f'generator-info-name="{xml_escape(GENERATOR_INFO_NAME)}">'
    )

    # --------------------------------------------------------
    # CHANNEL
    # --------------------------------------------------------

    valid_channels = []

    for channel in channels:

        channel_id = str(
            channel["id"]
        ).strip()

        ref = mapping.get(channel_id)

        if not ref:
            continue

        epg_channel = ref["channel"]
        display_name = ref["display-name"]

        if not epg_channel or not display_name:
            continue

        valid_channels.append({
            "id": channel_id,
            "channel": epg_channel,
            "display-name": display_name,
        })

        lines.append(
            f'  <channel id="{xml_escape(epg_channel)}">\n'
            f'    <display-name lang="vi">'
            f'{xml_escape(display_name)}'
            f'</display-name>\n'
            f'  </channel>'
        )

    # --------------------------------------------------------
    # PROGRAMME
    # --------------------------------------------------------

    for item in valid_channels:

        channel_id = item["id"]

        epg_channel = item["channel"]

        display_name = item["display-name"]

        schedules = epg_data.get(
            channel_id,
            []
        )

        for schedule in schedules:

            programme = build_programme(
                epg_channel,
                display_name,
                schedule
            )

            if programme:
                lines.append(programme.rstrip("\n"))

    lines.append("</tv>")

    return "\n".join(lines) + "\n"


# ============================================================
# LOG
# ============================================================

def load_previous_log_data():
    """
    Đọc trạng thái chạy trước từ một file JSON riêng.

    File này nằm trong tv360/:
        tv360_state.json

    File này không phải log hiển thị cho người dùng.
    """

    state_file = os.path.join(
        BASE_DIR,
        "tv360_state.json"
    )

    if not os.path.exists(state_file):
        return {
            "channels": []
        }

    try:

        with open(
            state_file,
            "r",
            encoding="utf-8"
        ) as f:

            return json.load(f)

    except Exception:

        return {
            "channels": []
        }


def save_current_state(channels):
    state_file = os.path.join(
        BASE_DIR,
        "tv360_state.json"
    )

    state = {
        "channels": [
            {
                "id": str(c["id"]),
                "name": c["name"],
            }
            for c in channels
        ]
    }

    with open(
        state_file,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            state,
            f,
            ensure_ascii=False,
            indent=2
        )


def update_log(
    channels,
    epg_data,
    mapping
):
    """
    Ghi log mới lên đầu.

    Giữ tối đa 7 bản ghi.
    """

    now = datetime.now()

    current_channels = {
        str(c["id"]): c["name"]
        for c in channels
    }

    previous_state = load_previous_log_data()

    previous_channels = {
        str(c["id"]): c["name"]
        for c in previous_state.get(
            "channels",
            []
        )
    }

    current_ids = set(
        current_channels.keys()
    )

    previous_ids = set(
        previous_channels.keys()
    )

    new_ids = current_ids - previous_ids

    removed_ids = previous_ids - current_ids

    # --------------------------------------------------------
    # EPG
    # --------------------------------------------------------

    epg_channels = []

    no_epg_channels = []

    for channel in channels:

        channel_id = str(
            channel["id"]
        )

        schedules = epg_data.get(
            channel_id,
            []
        )

        if schedules:
            epg_channels.append(channel)
        else:
            no_epg_channels.append(channel)

    # --------------------------------------------------------
    # Mapping
    # --------------------------------------------------------

    mapped = 0
    unmapped = []

    for channel in channels:

        channel_id = str(
            channel["id"]
        )

        ref = mapping.get(channel_id)

        if (
            ref
            and ref.get("channel")
            and ref.get("display-name")
        ):
            mapped += 1
        else:
            unmapped.append(channel)

    # --------------------------------------------------------
    # Log entry
    # --------------------------------------------------------

    lines = []

    lines.append(
        "=" * 70
    )

    lines.append(
        f"TV360 LOG - {now.strftime('%Y-%m-%d %H:%M:%S')}"
    )

    lines.append(
        "=" * 70
    )

    lines.append(
        f"Số lượng kênh hiện tại : {len(channels)}"
    )

    difference = (
        len(channels)
        - len(previous_channels)
    )

    if difference > 0:
        diff_text = f"+{difference}"

    else:
        diff_text = str(difference)

    lines.append(
        f"Chênh lệch so với lần trước : {diff_text}"
    )

    lines.append("")

    # --------------------------------------------------------
    # Kênh mới
    # --------------------------------------------------------

    lines.append(
        f"Kênh mới : {len(new_ids)}"
    )

    if new_ids:

        for channel_id in sorted(
            new_ids,
            key=lambda x: (
                int(x)
                if x.isdigit()
                else 999999999
            )
        ):

            lines.append(
                f"  + {channel_id} - "
                f"{current_channels[channel_id]}"
            )

    else:
        lines.append("  Không có")

    lines.append("")

    # --------------------------------------------------------
    # Kênh mất
    # --------------------------------------------------------

    lines.append(
        f"Kênh không còn : {len(removed_ids)}"
    )

    if removed_ids:

        for channel_id in sorted(
            removed_ids,
            key=lambda x: (
                int(x)
                if x.isdigit()
                else 999999999
            )
        ):

            lines.append(
                f"  - {channel_id} - "
                f"{previous_channels[channel_id]}"
            )

    else:
        lines.append("  Không có")

    lines.append("")

    # --------------------------------------------------------
    # EPG
    # --------------------------------------------------------

    lines.append(
        f"Kênh có EPG : {len(epg_channels)}"
    )

    lines.append(
        f"Kênh không có EPG : {len(no_epg_channels)}"
    )

    if no_epg_channels:

        for channel in no_epg_channels:

            lines.append(
                f"  ! {channel['id']} - "
                f"{channel['name']}"
            )

    else:

        lines.append("  Không có")

    lines.append("")

    # --------------------------------------------------------
    # Mapping
    # --------------------------------------------------------

    lines.append(
        f"Kênh đã mapping : {mapped}"
    )

    lines.append(
        f"Kênh chưa mapping : {len(unmapped)}"
    )

    if unmapped:

        for channel in unmapped:

            lines.append(
                f"  ? {channel['id']} - "
                f"{channel['name']}"
            )

    else:

        lines.append("  Không có")

    lines.append("")

    # --------------------------------------------------------
    # Đọc log cũ
    # --------------------------------------------------------

    old_entries = []

    if os.path.exists(LOG_FILE):

        with open(
            LOG_FILE,
            "r",
            encoding="utf-8"
        ) as f:

            old_text = f.read()

        # Mỗi entry bắt đầu bằng 70 dấu =
        parts = old_text.split(
            "=" * 70
        )

        for part in parts:

            part = part.strip()

            if part:
                old_entries.append(
                    part
                )

    new_entry = "\n".join(lines)

    entries = [
        new_entry
    ] + old_entries

    entries = entries[:MAX_LOG_DAYS]

    with open(
        LOG_FILE,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(
            (
                "\n\n"
                + "=" * 70
                + "\n\n"
            ).join(entries)
        )

    return {
        "current": len(channels),
        "difference": difference,
        "new": len(new_ids),
        "removed": len(removed_ids),
        "epg": len(epg_channels),
        "no_epg": len(no_epg_channels),
        "mapped": mapped,
        "unmapped": len(unmapped),
    }


# ============================================================
# MAIN
# ============================================================

def main():

    start_time = time.time()

    log_print("=" * 70)
    log_print("TV360 EPG START")
    log_print("=" * 70)

    # --------------------------------------------------------
    # 1. Lấy danh sách kênh
    # --------------------------------------------------------

    channels = get_all_channels()

    if not channels:

        raise RuntimeError(
            "Không lấy được danh sách kênh TV360."
        )

    # --------------------------------------------------------
    # 2. Excel
    # --------------------------------------------------------

    wb, ws, ws_ref = ensure_workbook()

    # Đọc mapping trước khi ghi Data
    mapping = read_reference_mapping(
        ws_ref
    )

    # Chỉ cập nhật A:E
    write_channels_to_excel(
        wb,
        ws,
        channels
    )

    log_print(
        "Đã cập nhật Data!A:E."
    )

    log_print(
        "Data!F:G được giữ nguyên."
    )

    # --------------------------------------------------------
    # 3. EPG
    # --------------------------------------------------------

    epg_data = {}

    total = len(channels)

    log_print(
        f"Bắt đầu lấy EPG cho {total} kênh..."
    )

    for index, channel in enumerate(
        channels,
        1
    ):

        channel_id = str(
            channel["id"]
        )

        log_print(
            f"[{index}/{total}] "
            f"{channel_id} - "
            f"{channel['name']}"
        )

        schedules = get_epg_for_channel(
            channel_id
        )

        epg_data[channel_id] = schedules

        log_print(
            f"    EPG: {len(schedules)} chương trình"
        )

        # Tránh request quá nhanh
        time.sleep(0.15)

    # --------------------------------------------------------
    # 4. XML
    # --------------------------------------------------------

    xml = build_xml(
        channels,
        epg_data,
        mapping
    )

    with open(
        XML_FILE,
        "w",
        encoding="utf-8",
        newline="\n"
    ) as f:

        f.write(xml)

    log_print(
        f"Đã tạo: {XML_FILE}"
    )

    # --------------------------------------------------------
    # 5. LOG
    # --------------------------------------------------------

    stats = update_log(
        channels,
        epg_data,
        mapping
    )

    # --------------------------------------------------------
    # 6. State
    # --------------------------------------------------------

    save_current_state(
        channels
    )

    # --------------------------------------------------------
    # DONE
    # --------------------------------------------------------

    elapsed = time.time() - start_time

    log_print("")
    log_print("=" * 70)
    log_print("TV360 EPG HOÀN TẤT")
    log_print("=" * 70)

    log_print(
        f"Kênh hiện tại : {stats['current']}"
    )

    log_print(
        f"Chênh lệch    : {stats['difference']:+d}"
    )

    log_print(
        f"Kênh mới      : {stats['new']}"
    )

    log_print(
        f"Kênh mất       : {stats['removed']}"
    )

    log_print(
        f"Có EPG        : {stats['epg']}"
    )

    log_print(
        f"Không có EPG  : {stats['no_epg']}"
    )

    log_print(
        f"Đã mapping     : {stats['mapped']}"
    )

    log_print(
        f"Chưa mapping   : {stats['unmapped']}"
    )

    log_print(
        f"Thời gian      : {elapsed:.1f} giây"
    )

    log_print("=" * 70)


if __name__ == "__main__":

    try:

        main()

    except Exception as exc:

        print("")
        print("=" * 70)
        print("TV360 ERROR")
        print("=" * 70)
        print(str(exc))
        traceback.print_exc()

        raise
