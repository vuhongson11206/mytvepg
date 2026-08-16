#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
MyTV EPG XML Generator
======================

Nguồn mapping:
    channel_list.xlsx

Cấu trúc cột:
    C = channel_id
    E = display-number
    F = XMLTV channel id
    G = display-name

API:
    Channel list:
        https://apigw.mytv.vn/api/v1/channel

    Schedule:
        https://apigw.mytv.vn/api/v1/channel/{channel_id}/schedule?date=YYYY-MM-DD

Logic:
    1. Đọc channel_list.xlsx.
    2. Mỗi dòng có channel_id ở cột C là một nguồn EPG độc lập.
    3. Các dòng có cùng F được gom thành cùng một XMLTV channel.
    4. Gọi API schedule cho TỪNG channel_id.
    5. Hợp nhất EPG của các channel_id cùng F.
    6. Loại chương trình trùng.
    7. Xuất XMLTV.
    8. EPG 3 ngày.
"""

import html
import os
import re
import sys
import time
from collections import OrderedDict
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook


# ============================================================
# CONFIG
# ============================================================

API_BASE_URL = (
    "https://apigw.mytv.vn/api/v1"
)

SCHEDULE_URL = (
    API_BASE_URL
    + "/channel/{channel_id}/schedule"
)

EXCEL_FILE = "channel_list.xlsx"

OUTPUT_FILE = "epg.xml"

# EPG 3 ngày:
# hôm nay + ngày mai + ngày kia
EPG_DAYS = 3

# Múi giờ XMLTV
XMLTV_TIMEZONE = "+0700"

# HTTP
REQUEST_TIMEOUT = 30
MAX_RETRIES = 3

# Delay giữa request
REQUEST_DELAY = 0.10

# Delay retry
RETRY_DELAY = 2

# Duration mặc định cho chương trình cuối
DEFAULT_PROGRAM_LENGTH = 30

# Không cho phép duration bất thường
MAX_PROGRAM_LENGTH = 24 * 60

USER_AGENT = (
    "Mozilla/5.0 "
    "(Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 "
    "(KHTML, like Gecko) "
    "Chrome/151.0 Safari/537.36"
)


# ============================================================
# EXCEL COLUMN INDEX
# ============================================================

# Excel:
#
# A = 1
# B = 2
# C = 3
# D = 4
# E = 5
# F = 6
# G = 7

COL_CHANNEL_ID = 3
COL_DISPLAY_NUMBER = 5
COL_XMLTV_CHANNEL_ID = 6
COL_DISPLAY_NAME = 7


# ============================================================
# HTTP SESSION
# ============================================================

session = requests.Session()

session.headers.update(
    {
        "User-Agent": USER_AGENT,
        "Accept": (
            "application/json, "
            "text/plain, */*"
        ),
        "Accept-Language": (
            "vi-VN,vi;q=0.9,"
            "en-US;q=0.8,en;q=0.7"
        ),
        "Connection": "keep-alive",
    }
)


# ============================================================
# LOG
# ============================================================

def log(message: str = "") -> None:
    print(message, flush=True)


def separator() -> None:
    log(
        "============================================================"
    )


# ============================================================
# STRING HELPERS
# ============================================================

def safe_string(value: Any) -> str:
    if value is None:
        return ""

    try:
        return str(value).strip()
    except Exception:
        return ""


def clean_text(value: Any) -> str:
    text = safe_string(value)

    if not text:
        return ""

    text = (
        text
        .replace("\r", " ")
        .replace("\n", " ")
        .replace("\t", " ")
    )

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


def xml_escape(value: Any) -> str:
    return html.escape(
        safe_string(value),
        quote=True
    )


def normalize_title(value: Any) -> str:
    """
    Chuẩn hóa title để phục vụ dedup.

    Ví dụ:
        "ABC News"
        " ABC   News "
        "abc news"

    sẽ được xem là cùng title.
    """

    text = clean_text(value)

    text = text.casefold()

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


def normalize_channel_id(value: Any) -> str:
    return clean_text(value)


# ============================================================
# HTTP JSON
# ============================================================

def get_json(
    url: str,
    description: str = ""
) -> Optional[Any]:

    last_error = None

    for attempt in range(
        1,
        MAX_RETRIES + 1
    ):

        try:

            log(
                f"GET {url} "
                f"(attempt {attempt}/{MAX_RETRIES})"
            )

            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT
            )

            response.raise_for_status()

            try:
                data = response.json()

            except ValueError as exc:

                last_error = exc

                log(
                    f"[JSON ERROR] "
                    f"{description}"
                )

                log(
                    response.text[:1000]
                )

                if attempt < MAX_RETRIES:

                    time.sleep(
                        RETRY_DELAY * attempt
                    )

                    continue

                return None

            if REQUEST_DELAY > 0:
                time.sleep(
                    REQUEST_DELAY
                )

            return data

        except requests.exceptions.Timeout as exc:

            last_error = exc

            log(
                f"[TIMEOUT] {description}"
            )

        except requests.exceptions.ConnectionError as exc:

            last_error = exc

            log(
                f"[CONNECTION ERROR] "
                f"{description}"
            )

        except requests.exceptions.HTTPError as exc:

            last_error = exc

            log(
                f"[HTTP ERROR] "
                f"{description}"
            )

            try:
                log(
                    f"HTTP {response.status_code}: "
                    f"{response.text[:500]}"
                )
            except Exception:
                pass

        except Exception as exc:

            last_error = exc

            log(
                f"[REQUEST ERROR] "
                f"{description}: "
                f"{type(exc).__name__}: {exc}"
            )

        if attempt < MAX_RETRIES:

            wait = (
                RETRY_DELAY * attempt
            )

            log(
                f"Retry sau {wait} giây..."
            )

            time.sleep(wait)

    log(
        f"[FAILED] {description}"
    )

    if last_error is not None:

        log(
            f"Last error: "
            f"{type(last_error).__name__}: "
            f"{last_error}"
        )

    return None


# ============================================================
# READ EXCEL
# ============================================================

def read_channel_list() -> List[Dict[str, Any]]:
    """
    Đọc channel_list.xlsx.

    Chỉ sử dụng:
        C = channel_id
        E = display-number
        F = XMLTV channel id
        G = display-name

    Không phụ thuộc tên header.
    """

    separator()

    log(
        f"ĐỌC FILE {EXCEL_FILE}"
    )

    separator()

    if not os.path.exists(
        EXCEL_FILE
    ):

        raise FileNotFoundError(
            f"Không tìm thấy "
            f"{EXCEL_FILE}"
        )

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True
    )

    try:

        # Lấy worksheet đầu tiên
        worksheet = workbook[
            workbook.sheetnames[0]
        ]

        log(
            f"Worksheet: "
            f"{worksheet.title}"
        )

        rows = []

        # ----------------------------------------------------
        # Bỏ qua dòng header
        # ----------------------------------------------------

        first_row = True

        for row_number, row in enumerate(
            worksheet.iter_rows(
                values_only=True
            ),
            start=1
        ):

            if first_row:

                first_row = False

                log(
                    "Bỏ qua dòng header:"
                )

                log(
                    f"  C={row[COL_CHANNEL_ID - 1]}"
                )

                log(
                    f"  E={row[COL_DISPLAY_NUMBER - 1]}"
                )

                log(
                    f"  F={row[COL_XMLTV_CHANNEL_ID - 1]}"
                )

                log(
                    f"  G={row[COL_DISPLAY_NAME - 1]}"
                )

                continue

            # ------------------------------------------------
            # Đảm bảo đủ cột
            # ------------------------------------------------

            if len(row) < COL_DISPLAY_NAME:

                continue

            channel_id = normalize_channel_id(
                row[
                    COL_CHANNEL_ID - 1
                ]
            )

            display_number = clean_text(
                row[
                    COL_DISPLAY_NUMBER - 1
                ]
            )

            xmltv_channel_id = clean_text(
                row[
                    COL_XMLTV_CHANNEL_ID - 1
                ]
            )

            display_name = clean_text(
                row[
                    COL_DISPLAY_NAME - 1
                ]
            )

            # ------------------------------------------------
            # Bỏ dòng không có channel_id
            # ------------------------------------------------

            if not channel_id:

                continue

            # ------------------------------------------------
            # F bắt buộc phải có
            # ------------------------------------------------

            if not xmltv_channel_id:

                log(
                    f"[SKIP ROW {row_number}] "
                    f"channel_id={channel_id} "
                    f"nhưng cột F trống."
                )

                continue

            # ------------------------------------------------
            # G nếu trống thì dùng F
            # ------------------------------------------------

            if not display_name:

                display_name = (
                    xmltv_channel_id
                )

            rows.append(
                {
                    "excel_row":
                        row_number,

                    "channel_id":
                        channel_id,

                    "display_number":
                        display_number,

                    "xmltv_id":
                        xmltv_channel_id,

                    "display_name":
                        display_name,
                }
            )

    finally:

        workbook.close()

    if not rows:

        raise RuntimeError(
            "Không tìm thấy dòng channel "
            "hợp lệ trong channel_list.xlsx."
        )

    log(
        f"Tổng số dòng channel hợp lệ: "
        f"{len(rows)}"
    )

    return rows


# ============================================================
# GROUP CHANNELS BY COLUMN F
# ============================================================

def group_channels(
    rows: List[Dict[str, Any]]
) -> OrderedDict:

    """
    GROUP BY cột F.

    Ví dụ:

        C=321 F=abcaustralia
        C=598 F=abcaustralia

    sẽ trở thành:

        abcaustralia:
            321
            598
    """

    groups = OrderedDict()

    seen_pair = set()

    for row in rows:

        xmltv_id = row[
            "xmltv_id"
        ]

        channel_id = row[
            "channel_id"
        ]

        pair = (
            xmltv_id,
            channel_id
        )

        # ----------------------------------------------------
        # Nếu cùng C xuất hiện nhiều lần cho cùng F,
        # chỉ lấy một lần.
        # ----------------------------------------------------

        if pair in seen_pair:

            log(
                f"[DUPLICATE EXCEL ROW] "
                f"F={xmltv_id} "
                f"C={channel_id} "
                f"-> bỏ dòng trùng."
            )

            continue

        seen_pair.add(pair)

        if xmltv_id not in groups:

            groups[
                xmltv_id
            ] = {
                "xmltv_id":
                    xmltv_id,

                "display_name":
                    row[
                        "display_name"
                    ],

                "display_number":
                    row[
                        "display_number"
                    ],

                "sources":
                    [],
            }

        group = groups[
            xmltv_id
        ]

        # ----------------------------------------------------
        # Nếu cùng F nhưng display-name khác nhau,
        # giữ giá trị của dòng đầu tiên.
        # ----------------------------------------------------

        if (
            group["display_name"]
            != row["display_name"]
        ):

            log(
                f"[WARNING] XMLTV ID "
                f"{xmltv_id} có display-name "
                f"khác nhau:"
            )

            log(
                f"  đang dùng: "
                f"{group['display_name']}"
            )

            log(
                f"  bỏ qua: "
                f"{row['display_name']}"
            )

        # ----------------------------------------------------
        # Tương tự display-number
        # ----------------------------------------------------

        if (
            group["display_number"]
            != row["display_number"]
        ):

            log(
                f"[WARNING] XMLTV ID "
                f"{xmltv_id} có display-number "
                f"khác nhau."
            )

        group[
            "sources"
        ].append(
            {
                "channel_id":
                    channel_id,

                "display_number":
                    row[
                        "display_number"
                    ],

                "excel_row":
                    row[
                        "excel_row"
                    ],
            }
        )

    return groups


# ============================================================
# PARSE API SCHEDULE LIST
# ============================================================

def extract_schedule_list(
    response: Any
) -> List[Any]:

    if response is None:
        return []

    if isinstance(
        response,
        list
    ):

        return response

    if not isinstance(
        response,
        dict
    ):

        return []

    # --------------------------------------------------------
    # data
    # --------------------------------------------------------

    data = response.get(
        "data"
    )

    if data is None:

        return []

    if isinstance(
        data,
        list
    ):

        return data

    if isinstance(
        data,
        dict
    ):

        # Trường hợp chuẩn:
        #
        # data:
        #   schedule: [...]

        for key in (
            "schedule",
            "programmes",
            "programs",
            "items",
            "list",
            "results",
        ):

            value = data.get(
                key
            )

            if isinstance(
                value,
                list
            ):

                return value

        # ----------------------------------------------------
        # Tìm list trong data
        # ----------------------------------------------------

        for value in data.values():

            if isinstance(
                value,
                list
            ):

                return value

    # --------------------------------------------------------
    # Một số response có schedule ở root
    # --------------------------------------------------------

    for key in (
        "schedule",
        "programmes",
        "programs",
        "items",
        "list",
        "results",
    ):

        value = response.get(
            key
        )

        if isinstance(
            value,
            list
        ):

            return value

    return []


# ============================================================
# API FIELD EXTRACTION
# ============================================================

def first_value(
    data: Dict[str, Any],
    keys: Iterable[str]
) -> Any:

    for key in keys:

        if key in data:

            value = data.get(
                key
            )

            if value is not None:

                return value

    return None


def extract_title(
    item: Dict[str, Any]
) -> str:

    value = first_value(
        item,
        [
            "title",
            "name",
            "program_name",
            "programName",
            "programme_name",
            "programmeName",
        ]
    )

    return clean_text(value)


def extract_time(
    item: Dict[str, Any]
) -> str:

    value = first_value(
        item,
        [
            "time",
            "start_time",
            "startTime",
            "begin_time",
            "beginTime",
            "start",
        ]
    )

    return clean_text(value)


def extract_description(
    item: Dict[str, Any]
) -> str:

    value = first_value(
        item,
        [
            "description",
            "desc",
            "summary",
            "content",
            "short_description",
            "shortDescription",
        ]
    )

    return clean_text(value)


def extract_length(
    item: Dict[str, Any]
) -> Optional[int]:

    value = first_value(
        item,
        [
            "duration",
            "duration_minutes",
            "durationMinutes",
            "length",
            "length_minutes",
            "lengthMinutes",
        ]
    )

    if value is None:
        return None

    try:

        number = float(
            value
        )

        if number > 0:

            return int(
                number
            )

    except (
        ValueError,
        TypeError
    ):

        pass

    text = safe_string(
        value
    )

    match = re.search(
        r"(\d+)",
        text
    )

    if match:

        try:

            result = int(
                match.group(1)
            )

            if result > 0:
                return result

        except ValueError:

            pass

    return None


# ============================================================
# PARSE DATETIME
# ============================================================

def parse_program_datetime(
    target_date: date,
    value: str
) -> Optional[datetime]:

    value = safe_string(
        value
    )

    if not value:
        return None

    # --------------------------------------------------------
    # HH:MM
    # --------------------------------------------------------

    match = re.match(
        r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$",
        value
    )

    if match:

        hour = int(
            match.group(1)
        )

        minute = int(
            match.group(2)
        )

        second = int(
            match.group(3) or 0
        )

        if not 0 <= hour <= 23:
            return None

        if not 0 <= minute <= 59:
            return None

        if not 0 <= second <= 59:
            return None

        return datetime(
            target_date.year,
            target_date.month,
            target_date.day,
            hour,
            minute,
            second
        )

    # --------------------------------------------------------
    # ISO datetime
    # --------------------------------------------------------

    iso_value = value

    if iso_value.endswith(
        "Z"
    ):

        iso_value = (
            iso_value[:-1]
            + "+00:00"
        )

    try:

        result = datetime.fromisoformat(
            iso_value
        )

        if result.tzinfo is not None:

            result = (
                result.astimezone()
                .replace(
                    tzinfo=None
                )
            )

        return result

    except ValueError:
        pass

    # --------------------------------------------------------
    # Các format phổ biến khác
    # --------------------------------------------------------

    formats = [
        "%H:%M",
        "%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
    ]

    for fmt in formats:

        try:

            result = datetime.strptime(
                value,
                fmt
            )

            # Nếu chỉ có giờ
            if result.year == 1900:

                result = datetime(
                    target_date.year,
                    target_date.month,
                    target_date.day,
                    result.hour,
                    result.minute,
                    result.second
                )

            return result

        except ValueError:

            continue

    return None


# ============================================================
# GET SCHEDULE FOR ONE CHANNEL / DATE
# ============================================================

def get_schedule(
    channel_id: str,
    target_date: date
) -> List[Dict[str, Any]]:

    date_string = (
        target_date.strftime(
            "%Y-%m-%d"
        )
    )

    url = (
        SCHEDULE_URL.format(
            channel_id=channel_id
        )
        + f"?date={date_string}"
    )

    response = get_json(
        url,
        (
            f"channel={channel_id} "
            f"date={date_string}"
        )
    )

    if response is None:

        log(
            f"[NO RESPONSE] "
            f"C={channel_id} "
            f"date={date_string}"
        )

        return []

    # --------------------------------------------------------
    # API trả data=None
    # --------------------------------------------------------

    if isinstance(
        response,
        dict
    ):

        if (
            "data" in response
            and response.get("data") is None
        ):

            log(
                f"[NO DATA] "
                f"C={channel_id} "
                f"date={date_string}"
            )

            return []

    raw_programmes = (
        extract_schedule_list(
            response
        )
    )

    if not raw_programmes:

        log(
            f"[NO EPG] "
            f"C={channel_id} "
            f"date={date_string}"
        )

        return []

    programmes = []

    for item in raw_programmes:

        if not isinstance(
            item,
            dict
        ):

            continue

        title = extract_title(
            item
        )

        if not title:

            title = "Chương trình"

        time_value = extract_time(
            item
        )

        start = (
            parse_program_datetime(
                target_date,
                time_value
            )
        )

        if start is None:

            log(
                f"[INVALID TIME] "
                f"C={channel_id} "
                f"date={date_string} "
                f"time={time_value} "
                f"title={title}"
            )

            continue

        programmes.append(
            {
                "title":
                    title,

                "start":
                    start,

                "description":
                    extract_description(
                        item
                    ),

                "api_length":
                    extract_length(
                        item
                    ),

                "source_channel_id":
                    channel_id,

                "raw":
                    item,
            }
        )

    programmes.sort(
        key=lambda item:
        item["start"]
    )

    log(
        f"[EPG OK] "
        f"C={channel_id} "
        f"date={date_string} "
        f"programmes={len(programmes)}"
    )

    return programmes


# ============================================================
# DEDUP
# ============================================================

def programme_dedup_key(
    programme: Dict[str, Any]
) -> Tuple:

    """
    Khóa chống duplicate.

    Ưu tiên:
        start
        normalized title

    Stop không đưa vào khóa vì hai nguồn có thể
    tính stop hơi khác nhau.

    Ví dụ:

        10:00 ABC News

    từ channel 321 và 598
    -> chỉ giữ 1.
    """

    return (
        programme[
            "start"
        ],
        normalize_title(
            programme[
                "title"
            ]
        ),
    )


def merge_programmes(
    programmes: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:

    """
    Hợp nhất EPG của nhiều channel_id.

    Nếu trùng:
        start + title

    thì chỉ giữ một chương trình.

    Nếu hai bản trùng nhau nhưng một bản có
    description/duration đầy đủ hơn, ưu tiên
    bản đầy đủ hơn.
    """

    merged = OrderedDict()

    for programme in programmes:

        key = programme_dedup_key(
            programme
        )

        if key not in merged:

            merged[
                key
            ] = programme

            continue

        current = merged[
            key
        ]

        # ----------------------------------------------------
        # Ưu tiên description đầy đủ
        # ----------------------------------------------------

        current_description = clean_text(
            current.get(
                "description",
                ""
            )
        )

        new_description = clean_text(
            programme.get(
                "description",
                ""
            )
        )

        if (
            len(new_description)
            > len(current_description)
        ):

            current[
                "description"
            ] = new_description

        # ----------------------------------------------------
        # Ưu tiên api_length hợp lệ
        # ----------------------------------------------------

        current_length = current.get(
            "api_length"
        )

        new_length = programme.get(
            "api_length"
        )

        if (
            not current_length
            and new_length
        ):

            current[
                "api_length"
            ] = new_length

    result = list(
        merged.values()
    )

    result.sort(
        key=lambda item:
        item["start"]
    )

    return result


# ============================================================
# CALCULATE STOP
# ============================================================

def calculate_stops(
    programmes: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:

    result = []

    for index, programme in enumerate(
        programmes
    ):

        start = programme[
            "start"
        ]

        # ----------------------------------------------------
        # Chương trình kế tiếp
        # ----------------------------------------------------

        if index + 1 < len(
            programmes
        ):

            next_start = (
                programmes[
                    index + 1
                ][
                    "start"
                ]
            )

            seconds = (
                next_start - start
            ).total_seconds()

            length = int(
                seconds / 60
            )

            # Nếu thời gian kế tiếp không hợp lệ
            if length <= 0:

                length = (
                    programme.get(
                        "api_length"
                    )
                    or DEFAULT_PROGRAM_LENGTH
                )

        else:

            length = (
                programme.get(
                    "api_length"
                )
                or DEFAULT_PROGRAM_LENGTH
            )

        # ----------------------------------------------------
        # Giới hạn duration
        # ----------------------------------------------------

        if length <= 0:

            length = (
                DEFAULT_PROGRAM_LENGTH
            )

        if length > MAX_PROGRAM_LENGTH:

            length = (
                DEFAULT_PROGRAM_LENGTH
            )

        stop = (
            start
            + timedelta(
                minutes=length
            )
        )

        item = dict(
            programme
        )

        item[
            "length"
        ] = length

        item[
            "stop"
        ] = stop

        result.append(
            item
        )

    return result


# ============================================================
# XMLTV DATETIME
# ============================================================

def xmltv_datetime(
    value: datetime
) -> str:

    return value.strftime(
        "%Y%m%d%H%M%S "
        + XMLTV_TIMEZONE
    )


# ============================================================
# BUILD CHANNEL XML
# ============================================================

def build_channel_xml(
    group: Dict[str, Any]
) -> List[str]:

    lines = []

    xmltv_id = xml_escape(
        group[
            "xmltv_id"
        ]
    )

    display_name = xml_escape(
        group[
            "display_name"
        ]
    )

    display_number = xml_escape(
        group[
            "display_number"
        ]
    )

    lines.append(
        f'  <channel id="{xmltv_id}">'
    )

    lines.append(
        f'    <display-name lang="vi">'
        f'{display_name}'
        f'</display-name>'
    )

    if display_number:

        lines.append(
            f'    <display-number>'
            f'{display_number}'
            f'</display-number>'
        )

    lines.append(
        "  </channel>"
    )

    return lines


# ============================================================
# BUILD PROGRAMME XML
# ============================================================

def build_programme_xml(
    xmltv_id: str,
    programme: Dict[str, Any]
) -> List[str]:

    lines = []

    channel = xml_escape(
        xmltv_id
    )

    start = xmltv_datetime(
        programme[
            "start"
        ]
    )

    stop = xmltv_datetime(
        programme[
            "stop"
        ]
    )

    title = xml_escape(
        programme[
            "title"
        ]
    )

    length = int(
        programme[
            "length"
        ]
    )

    description = clean_text(
        programme.get(
            "description",
            ""
        )
    )

    # --------------------------------------------------------
    # Yêu cầu:
    #
    # Chương trình này có thời lượng X phút
    # --------------------------------------------------------

    duration_text = (
        f"Chương trình này có thời lượng "
        f"{length} phút"
    )

    if description:

        full_description = (
            f"{description} "
            f"{duration_text}"
        )

    else:

        full_description = (
            duration_text
        )

    full_description = xml_escape(
        full_description
    )

    lines.append(
        f'  <programme '
        f'start="{start}" '
        f'stop="{stop}" '
        f'channel="{channel}">'
    )

    lines.append(
        f'    <title lang="vi">'
        f'{title}'
        f'</title>'
    )

    lines.append(
        f'    <desc lang="vi">'
        f'{full_description}'
        f'</desc>'
    )

    lines.append(
        f'    <length lang="vi">'
        f'Chương trình này có thời lượng '
        f'{length} phút'
        f'</length>'
    )

    lines.append(
        "  </programme>"
    )

    return lines


# ============================================================
# PROCESS ONE XMLTV CHANNEL GROUP
# ============================================================

def process_group(
    group: Dict[str, Any],
    target_date: date
) -> List[Dict[str, Any]]:

    xmltv_id = group[
        "xmltv_id"
    ]

    sources = group[
        "sources"
    ]

    log("")
    log(
        f"XMLTV CHANNEL: "
        f"{xmltv_id}"
    )

    log(
        f"Display name: "
        f"{group['display_name']}"
    )

    log(
        f"Sources: "
        f"{len(sources)} channel_id"
    )

    all_programmes = []

    # --------------------------------------------------------
    # Lấy EPG từng channel_id
    # --------------------------------------------------------

    for source in sources:

        channel_id = source[
            "channel_id"
        ]

        log(
            f"  -> API C={channel_id}"
        )

        programmes = get_schedule(
            channel_id,
            target_date
        )

        all_programmes.extend(
            programmes
        )

    # --------------------------------------------------------
    # Merge + dedup
    # --------------------------------------------------------

    merged = merge_programmes(
        all_programmes
    )

    merged = calculate_stops(
        merged
    )

    log(
        f"MERGED: "
        f"{len(all_programmes)} "
        f"raw -> "
        f"{len(merged)} unique"
    )

    return merged


# ============================================================
# CREATE EPG
# ============================================================

def create_epg(
    groups: OrderedDict
) -> Tuple[int, int, int]:

    today = date.today()

    xml_lines = []

    total_raw = 0
    total_unique = 0
    failed_days = 0

    # ========================================================
    # XML HEADER
    # ========================================================

    xml_lines.append(
        '<?xml version="1.0" encoding="UTF-8"?>'
    )

    xml_lines.append(
        '<tv '
        'source-info-name="MyTV" '
        'source-info-url="https://mytv.com.vn" '
        'generator-info-name="MyTV EPG GitHub"'
        '>'
    )

    # ========================================================
    # CHANNELS
    # ========================================================

    separator()

    log(
        "TẠO CHANNEL XML"
    )

    separator()

    for group in groups.values():

        xml_lines.extend(
            build_channel_xml(
                group
            )
        )

    # ========================================================
    # EPG
    # ========================================================

    for day_offset in range(
        EPG_DAYS
    ):

        target_date = (
            today
            + timedelta(
                days=day_offset
            )
        )

        log("")
        separator()

        log(
            f"EPG NGÀY "
            f"{day_offset + 1}/"
            f"{EPG_DAYS}: "
            f"{target_date}"
        )

        separator()

        for group_index, group in enumerate(
            groups.values(),
            start=1
        ):

            log(
                f"[{group_index}/"
                f"{len(groups)}] "
                f"{group['xmltv_id']} "
                f"- "
                f"{group['display_name']}"
            )

            try:

                programmes = process_group(
                    group,
                    target_date
                )

            except Exception as exc:

                failed_days += 1

                log(
                    f"[GROUP ERROR] "
                    f"{group['xmltv_id']} "
                    f"{target_date}"
                )

                log(
                    f"{type(exc).__name__}: "
                    f"{exc}"
                )

                continue

            total_unique += len(
                programmes
            )

            for programme in programmes:

                xml_lines.extend(
                    build_programme_xml(
                        group[
                            "xmltv_id"
                        ],
                        programme
                    )
                )

    # ========================================================
    # XML END
    # ========================================================

    xml_lines.append(
        "</tv>"
    )

    xml_content = (
        "\n".join(
            xml_lines
        )
        + "\n"
    )

    # ========================================================
    # WRITE FILE
    # ========================================================

    log("")
    separator()

    log(
        f"GHI {OUTPUT_FILE}"
    )

    separator()

    with open(
        OUTPUT_FILE,
        "w",
        encoding="utf-8",
        newline="\n"
    ) as file:

        file.write(
            xml_content
        )

    return (
        total_raw,
        total_unique,
        failed_days
    )


# ============================================================
# VALIDATE XML
# ============================================================

def validate_xml() -> bool:

    if not os.path.exists(
        OUTPUT_FILE
    ):

        log(
            "[XML ERROR] "
            "Không tồn tại epg.xml"
        )

        return False

    try:

        import xml.etree.ElementTree as ET

        tree = ET.parse(
            OUTPUT_FILE
        )

        root = tree.getroot()

        if root.tag != "tv":

            log(
                "[XML ERROR] "
                f"Root là <{root.tag}> "
                f"thay vì <tv>."
            )

            return False

        channels = root.findall(
            "channel"
        )

        programmes = root.findall(
            "programme"
        )

        # ----------------------------------------------------
        # Kiểm tra channel ID duplicate
        # ----------------------------------------------------

        channel_ids = []

        for channel in channels:

            channel_id = (
                channel.get(
                    "id"
                )
            )

            if channel_id:
                channel_ids.append(
                    channel_id
                )

        duplicates = (
            len(channel_ids)
            != len(set(channel_ids))
        )

        if duplicates:

            log(
                "[XML ERROR] "
                "Có channel id bị duplicate."
            )

            return False

        # ----------------------------------------------------
        # Kiểm tra programme channel
        # ----------------------------------------------------

        valid_channel_ids = set(
            channel_ids
        )

        invalid_programmes = 0

        for programme in programmes:

            channel_id = (
                programme.get(
                    "channel"
                )
            )

            if (
                channel_id
                not in valid_channel_ids
            ):

                invalid_programmes += 1

        if invalid_programmes:

            log(
                f"[XML WARNING] "
                f"{invalid_programmes} programme "
                f"tham chiếu channel không tồn tại."
            )

        # ----------------------------------------------------
        # Summary
        # ----------------------------------------------------

        size = os.path.getsize(
            OUTPUT_FILE
        )

        separator()

        log(
            "XML VALIDATION"
        )

        separator()

        log(
            f"Channels:   "
            f"{len(channels)}"
        )

        log(
            f"Programmes: "
            f"{len(programmes)}"
        )

        log(
            f"File size:  "
            f"{size:,} bytes"
        )

        separator()

        return True

    except Exception as exc:

        log(
            "[XML ERROR] "
            f"{type(exc).__name__}: "
            f"{exc}"
        )

        return False


# ============================================================
# MAIN
# ============================================================

def main() -> int:

    started = time.time()

    separator()

    log(
        "              MyTV EPG"
    )

    log(
        "       XMLTV Generator"
    )

    log(
        f"       EPG {EPG_DAYS} DAYS"
    )

    separator()

    log(
        f"Excel:       {EXCEL_FILE}"
    )

    log(
        f"Output:      {OUTPUT_FILE}"
    )

    log(
        f"EPG days:    {EPG_DAYS}"
    )

    log(
        f"Timezone:    {XMLTV_TIMEZONE}"
    )

    log("")

    # ========================================================
    # STEP 1: READ EXCEL
    # ========================================================

    try:

        rows = read_channel_list()

    except Exception as exc:

        separator()

        log(
            "FATAL ERROR - EXCEL"
        )

        log(
            f"{type(exc).__name__}: "
            f"{exc}"
        )

        separator()

        return 1

    # ========================================================
    # STEP 2: GROUP BY F
    # ========================================================

    separator()

    log(
        "GROUP CHANNEL THEO CỘT F"
    )

    separator()

    groups = group_channels(
        rows
    )

    log(
        f"Tổng dòng Excel: "
        f"{len(rows)}"
    )

    log(
        f"Tổng XMLTV channel: "
        f"{len(groups)}"
    )

    duplicate_source_count = (
        sum(
            max(
                0,
                len(
                    group["sources"]
                ) - 1
            )
            for group in groups.values()
        )
    )

    log(
        f"Nguồn channel_id bổ sung "
        f"do F trùng: "
        f"{duplicate_source_count}"
    )

    for group in groups.values():

        source_ids = ",".join(
            source["channel_id"]
            for source in group["sources"]
        )

        log(
            f"  {group['xmltv_id']} "
            f"| "
            f"{group['display_name']} "
            f"| C={source_ids}"
        )

    # ========================================================
    # STEP 3: CREATE EPG
    # ========================================================

    try:

        (
            total_raw,
            total_unique,
            failed_days
        ) = create_epg(
            groups
        )

    except Exception as exc:

        separator()

        log(
            "FATAL ERROR - EPG"
        )

        log(
            f"{type(exc).__name__}: "
            f"{exc}"
        )

        separator()

        return 1

    # ========================================================
    # STEP 4: VALIDATE
    # ========================================================

    valid = validate_xml()

    # ========================================================
    # SUMMARY
    # ========================================================

    elapsed = (
        time.time()
        - started
    )

    separator()

    log(
        "                  SUMMARY"
    )

    separator()

    log(
        f"Excel rows:        {len(rows)}"
    )

    log(
        f"XMLTV channels:    {len(groups)}"
    )

    log(
        f"EPG days:          {EPG_DAYS}"
    )

    log(
        f"Unique programmes: {total_unique}"
    )

    log(
        f"Group errors:      {failed_days}"
    )

    log(
        f"XML valid:         "
        f"{'YES' if valid else 'NO'}"
    )

    log(
        f"Output:            {OUTPUT_FILE}"
    )

    log(
        f"Elapsed:           "
        f"{elapsed:.2f} seconds"
    )

    separator()

    if not valid:

        return 1

    if total_unique == 0:

        log(
            "[ERROR] "
            "Không tạo được chương trình EPG."
        )

        return 1

    return 0


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    sys.exit(
        main()
    )
